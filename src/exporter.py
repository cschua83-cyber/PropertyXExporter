import os

from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)

HEADERS = [
    "Phase",
    "Block",
    "Unit",
    "Bedroom",
    "Bathroom",
    "Built-up",
    "List Price\nRM",
    "SPA Price\nRM",
    "Net Price\nRM",
    "PSF\nRM",
    "Direction",
    "Car Park",
]

HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

def format_header(ws):

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.row_dimensions[1].height = 35
    

def format_sheet(ws):

    ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT
            

def format_numbers(ws):

    for column in ["G", "H", "I"]:
        for cell in ws[column][1:]:
            cell.number_format = '#,##0;-#,##0'

    for cell in ws["J"][1:]:
        cell.number_format = '#,##0'
        

def auto_fit_columns(ws):

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2


def auto_fit_columns(ws):

    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2
            

def export_to_excel(units):
    
    os.makedirs(
    "output",
    exist_ok=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Units"

    ws.append(HEADERS)

    format_header(ws)    
        
    for u in units:
         
        ws.append([
            u.phase,
            u.block,
            u.unit,
            u.bedroom,
            u.bathroom,
            u.size,
            u.list_price,      
            u.spa_price,
            u.net_price,
            u.psf,
            u.orientation,
            u.carpark,
        ])

    format_sheet(ws)

    format_numbers(ws)

    filename = f"output/Units_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    
    auto_fit_columns(ws)

    wb.save(filename)

    print(f"✅ Excel 已输出：{filename}")